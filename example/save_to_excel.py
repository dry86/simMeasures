import pandas as pd
from openpyxl import load_workbook
import os

# 定义保存到Excel的函数，支持指定工作表
def save_to_excel(cal_method, score, row, sheet, file_name="/newdisk/public/wws/simMeasures/example/results.xlsx"):
    # 如果文件不存在，创建新文件并写入数据
    if not os.path.exists(file_name):
        # 创建一个新的 DataFrame，并保存到指定工作表中
        df = pd.DataFrame({cal_method: [None] * row})
        df.loc[row - 1, cal_method] = score
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    else:
        # 文件已存在，加载现有文件
        book = load_workbook(file_name)
        
        # 检查工作表是否存在
        if sheet in book.sheetnames:
            df_existing = pd.read_excel(file_name, sheet_name=sheet)
        else:
            # 如果工作表不存在，创建一个新的空DataFrame
            df_existing = pd.DataFrame()

        # 如果列已经存在，则追加数据
        if cal_method in df_existing.columns:
            if len(df_existing) < row:  # 扩展行数
                df_existing = df_existing.reindex(list(range(row)))
            df_existing.loc[row - 1, cal_method] = score
        else:
            # 如果列不存在，创建新列并填充数据
            df_existing[cal_method] = [None] * len(df_existing)
            if len(df_existing) < row:  # 同样检查是否需要扩展行数
                df_existing = df_existing.reindex(list(range(row)))
            df_existing.loc[row - 1, cal_method] = score
        
        # 使用 openpyxl 引擎以附加模式写入工作簿，而不会覆盖其他工作表
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # writer.book = book  # 使用已加载的工作簿
            df_existing.to_excel(writer, sheet_name=sheet, index=False)

def save_to_excel_2(cal_method, score, row, sheet, file_name="/newdisk/public/wws/simMeasures/results/codellama_7b_and_7b_python/test.xlsx"):

    # 如果文件不存在，创建新文件并写入数据
    if not os.path.exists(file_name):
        # 创建一个新的 DataFrame，并保存到指定工作表中
        df = pd.DataFrame({cal_method: [None] * row})
        df.loc[row - 1, cal_method] = score
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    else:
        # 文件已存在，加载现有文件
        book = load_workbook(file_name)
        
        # 检查工作表是否存在
        if sheet in book.sheetnames:
            df_existing = pd.read_excel(file_name, sheet_name=sheet)
        else:
            # 如果工作表不存在，创建一个新的空DataFrame
            df_existing = pd.DataFrame()

        # 如果列已经存在，则追加数据
        if cal_method in df_existing.columns:
            if len(df_existing) < row:  # 扩展行数
                df_existing = df_existing.reindex(list(range(row)))
            df_existing.loc[row - 1, cal_method] = score
        else:
            # 如果列不存在，创建新列并填充数据
            df_existing[cal_method] = [None] * len(df_existing)
            if len(df_existing) < row:  # 同样检查是否需要扩展行数
                df_existing = df_existing.reindex(list(range(row)))
            df_existing.loc[row - 1, cal_method] = score
        
        # 使用 openpyxl 引擎以附加模式写入工作簿，而不会覆盖其他工作表
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # writer.book = book  # 使用已加载的工作簿
            df_existing.to_excel(writer, sheet_name=sheet, index=False)

        # 重新打开文件，设置单元格格式为数字类型，避免科学计数法
        book = load_workbook(file_name)
        sheet_to_format = book[sheet]
        
        # 设置单元格格式为浮点型数字显示
        for row_cells in sheet_to_format.iter_rows(min_row=2, max_row=row, min_col=1, max_col=len(df_existing.columns)):
            for cell in row_cells:
                if isinstance(cell.value, float):  # 只对小数类型的单元格进行格式设置
                    cell.number_format = '0.0000000000000000'  # 设置足够的位数显示小数点后的所有位数

        # 保存工作簿
        book.save(file_name)

# 打印并保存计算结果的函数
def print_and_save(cal_method, score, row, sheet):
    # 打印结果
    print(f"\t {cal_method}: {score}")
    
    # 调用函数保存数据到指定工作表中
    save_to_excel_2(cal_method, score, row, sheet)


if __name__ == "__main__":
    # 示例使用
    pwcca_mean = 0.789456123456789  # 这是一个示例值
    print_and_save("PWCCA similarity", pwcca_mean, row=1, sheet="Metrics")

    pwcca_mean = 0.897656342134534116
    print_and_save("PWCCA similarity", pwcca_mean, row=2, sheet="Metrics")

    # 示例添加另一个cal_method
    cosine_sim = 0.78
    print_and_save("Cosine similarity", cosine_sim, row=1, sheet="Metrics")

    print_and_save("Cosine similarity", cosine_sim, row=1, sheet="Metrics1")