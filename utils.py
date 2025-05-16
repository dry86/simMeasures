import os
import torch
import pandas as pd
import collections
import math
import json
from PIL import Image
from openpyxl import load_workbook
from typing import List, Tuple, Optional, Dict, Any
from transformers import AutoTokenizer, AutoModelForCausalLM
from abc import ABC, abstractmethod


class VLModel(ABC):
    def __init__(self, model_path, device):
        self.model_path = model_path
        self.device = device

    @abstractmethod
    def load_model(self):
        pass

    @abstractmethod
    def get_hidden_states(self, image_path: str, text: str):
        pass

class QwenVLModel(VLModel):
    def load_model(self):
        from transformers import Qwen2VLForConditionalGeneration, AutoProcessor
        self.model = Qwen2VLForConditionalGeneration.from_pretrained(
            self.model_path, torch_dtype="auto"
        ).to(self.device).eval()
        self.processor = AutoProcessor.from_pretrained(self.model_path)

    def get_hidden_states(self, image_path: str, text: str):
        image = Image.open(image_path)
        conversation = [
            {
                "role": "user",
                "content": [{"type": "image"}, {"type": "text", "text": text}],
            }
        ]
        prompt = self.processor.apply_chat_template(conversation, add_generation_prompt=True)
        inputs = self.processor(text=[prompt], images=[image], return_tensors="pt").to(self.device)

        with torch.no_grad():
            outputs = self.model(**inputs, output_hidden_states=True, return_dict=True)

        last_non_pad_idx = inputs['attention_mask'].sum(dim=1) - 1
        hidden_states = [
            layer[0, last_non_pad_idx, :].squeeze().cpu() for layer in outputs.hidden_states
        ]
        return hidden_states

MODEL_REGISTRY = {
    "qwen_vl": QwenVLModel,
    # "blip2": Blip2Model,
    # 添加更多模型名和类的映射
}

def get_model_class(model_type: str):
    if model_type not in MODEL_REGISTRY:
        raise ValueError(f"Unsupported model type: {model_type}")
    return MODEL_REGISTRY[model_type]


def _extract_from_jsonl(
    file_path: str,
    key: Optional[str],
    field: str,
    prefix: str
) -> List[Tuple[Optional[str], str]]:
    """从 jsonl 文件中提取 (task_id, prompt) 的列表。"""
    import jsonlines
    prompts = []
    with jsonlines.open(file_path) as reader:
        for obj in reader:
            task_id = obj.get(key) if key else None
            content = obj.get(field, "")
            prompt = prefix + content
            prompts.append((task_id, prompt))
    return prompts

def _extract_from_parquet(
    file_path: str,
    key: Optional[str],
    field: str,
    prefix: str
) -> List[Tuple[Optional[str], str]]:
    """从 parquet 文件中提取 (task_id, prompt) 的列表。"""
    prompts = []
    df = pd.read_parquet(file_path)
    for idx, row in df.iterrows():
        task_id = row.get(key) if key else None
        content = row.get(field, "")
        prompt = prefix + content
        prompts.append((task_id, prompt))
    return prompts

def _extract_from_txt(
    file_path: str,
    prefix: str
) -> List[Tuple[Optional[int], str]]:
    """从纯文本文件（逐行）提取 (line_number, prompt) 的列表。"""
    prompts = []
    with open(file_path, "r", encoding="utf-8") as file:
        lines = [line.strip() for line in file]
    for i, line in enumerate(lines, start=1):
        prompt = prefix + line
        prompts.append((i, prompt))
    return prompts

def _extract_from_json(
    file_path: str,
    key: Optional[str],
    field: Optional[str],
    prefix: str
) -> List[Tuple[Optional[Any], str]]:
    """
    从普通 JSON 文件中提取 (task_id, prompt) 的列表。
    假设该 JSON 文件的顶层数据是一个列表，每个元素是一个 dict。
    """
    import json
    prompts = []
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)  # 假设 data 是一个 list[dict]
        if not isinstance(data, list):
            raise ValueError(f"Expected a list of objects in JSON file: {file_path}")
        for obj in data:
            task_id = obj.get(key) if key else None
            content = obj.get(field, "")
            prompt = prefix + content
            prompts.append((task_id, prompt))
    return prompts

def extract_prompts(mode: str, lang: str) -> List[Tuple[Optional[Any], str]]:
    """
    通用函数，用于从文件中提取Prompts。

    参数:
        mode (str): 解析模式，可选值：
            - 'textGen_MBPP': 提取'text'或'prompt'字段。
            - 'textGen_humaneval': 提取'prompt'字段。
            - 'codeSummary_CSearchNet': 提取'code'字段，并添加固定前缀。
            - 'codeRepair': 从纯文本文件逐行读取，并添加固定前缀。
            - 'line_completion': 提取'input'字段。

        lang (str): 语言标签，例如 'python', 'java' 等。

    返回:
        List[Tuple[Optional[Any], str]]: 提取的Prompt列表。每个元素为 (task_id或其他标识, prompt)。
    """

    # 将每种模式的具体配置放在 mode_config 中
    # file_type: 用于区分使用哪个小函数去读取
    # path: 文件路径，可以包含 {lang} 占位符
    # field: 提取的字段（jsonl/parquet）
    # prefix: 拼接前缀
    # key: 在数据中用来标识任务ID的字段名；若不需要可以为 None

    mode_config: Dict[str, Dict[str, Any]] = {
        'textGen_MBPP': {
            'key': 'task_id',
            'path_template': [
                {
                    'file_type': 'jsonl',
                    'path': "/newdisk/public/wws/Dataset/mbpp/{lang}/mbpp.jsonl",
                    'field': 'text',
                    'prefix': ""
                },
                {
                    'file_type': 'parquet',
                    'path': "/newdisk/public/wws/Dataset/mbpp/data_MultiPL-E/{lang}/test-00000-of-00001.parquet",
                    'field': 'prompt',
                    'prefix': ""
                }
            ]
        },
        'textGen_humaneval': {
            'key': 'task_id',
            'path_template': [
                {
                    'file_type': 'jsonl',
                    'path': "/newdisk/public/wws/Dataset/humaneval-x-main/data/{lang}/data/humaneval.jsonl",
                    'field': 'prompt',
                    'prefix': ""
                },
                {
                    'file_type': 'parquet',
                    'path': "/newdisk/public/wws/Dataset/humaneval-x-main/data_MultiPL-E/{lang}/test-00000-of-00001.parquet",
                    'field': 'prompt',
                    'prefix': ""
                }
            ]
        },
        'codeSummary_CSearchNet': {
            'key': 'repo',
            'path_template': [
                {
                    'file_type': 'jsonl',
                    'path': "/newdisk/public/wws/Dataset/CodeSearchNet/dataset/{lang}/test.jsonl",
                    'field': 'code',
                    'prefix': "Please describe the functionality of the method: "
                }
            ]
        },
        'codeRepair': {
            'key': None,
            'path_template': [
                {
                    'file_type': 'txt',
                    'path': "/newdisk/public/wws/Dataset/code-refinement/data/small/test.buggy-fixed.buggy",
                    'field': None,  # 不适用
                    'prefix': "Please fix the bug in the following code: "
                }
            ]
        },
        'line_completion': {
            'key': 'id',
            'path_template': [
                {
                    'file_type': 'json',
                    'path': "/newdisk/public/wws/Dataset/CodeCompletion-line/dataset/{lang}/line_completion/test.json",
                    'field': 'input',
                    'prefix': ""
                }
            ]
        }
    }

    if mode not in mode_config:
        raise ValueError(
            f"Invalid Task mode: {mode}. Available modes: {list(mode_config.keys())}"
        )

    config = mode_config[mode]
    all_prompts: List[Tuple[Optional[Any], str]] = []

    # 遍历当前模式下的多个文件配置
    for path_conf in config['path_template']:
        file_type: str = path_conf['file_type']
        file_path: str = path_conf['path'].format(lang=lang)
        key: Optional[str] = config.get('key', None)  # mode-level key
        field: Optional[str] = path_conf.get('field', None)
        prefix: str = path_conf.get('prefix', "")

        if not os.path.exists(file_path):
            print(f"File not found, skip: {file_path}")
            continue

        print(f"Processing file: {file_path}")

        # 根据 file_type 选择合适的提取函数
        if file_type == 'jsonl':
            if not field:
                raise ValueError("field must be provided for JSONL file.")
            # 提取并累加
            prompts = _extract_from_jsonl(file_path, key, field, prefix)
            all_prompts.extend(prompts)

        elif file_type == 'parquet':
            if not field:
                raise ValueError("field must be provided for Parquet file.")
            prompts = _extract_from_parquet(file_path, key, field, prefix)
            all_prompts.extend(prompts)

        elif file_type == 'txt':
            # codeRepair 这种场景：逐行提取
            prompts = _extract_from_txt(file_path, prefix)
            all_prompts.extend(prompts)

        elif file_type == 'json':
            prompts = _extract_from_jsonl(file_path, key, field, prefix)
            all_prompts.extend(prompts)

        else:
            print(f"Unknown file_type '{file_type}' in config, skipping: {file_path}")

    return all_prompts


class BaseResultSaver(ABC):
    """
    抽象基类，定义 print_and_save 接口和 save_result 接口。
    子类需实现 save_result 方法。
    """
    def __init__(self, file_path, sheet=None, model1=None, model2=None):
        self.file_path = file_path
        self.sheet = sheet  # 对于 JSONL/Parquet 可忽略或不使用
        # 在初始化时对 (model1, model2) 做排序，存成 self.m1, self.m2
        if model1 and model2:
            self.m1, self.m2 = self.canonical_pair(model1, model2)
        else:
            self.m1, self.m2 = None, None

    def print_and_save(self, cal_method, score, row):
        # 统一的打印逻辑
        print(f"\t{cal_method}: {score}")
        # 调用子类实现的保存逻辑
        self.save_result(cal_method, score, row)

    @abstractmethod
    def print_and_save(self, metrics_dict, row):
        pass

    # @abstractmethod
    # def save_result(self, cal_method, score, row):
    #     """
    #     子类需实现的保存逻辑
    #     """
    #     pass

    def canonical_pair(self, m1, m2):
        return tuple(sorted([m1, m2]))


class ExcelResultSaver(BaseResultSaver):
    """
    将数据追加到 Excel 文件中的指定工作表。
        如果文件不存在，创建新文件并写入数据。
        如果文件已存在，将数据追加到指定工作表。
    """
    def __init__(self, file_path, sheet="Sheet1"):
        super().__init__(file_path, sheet=sheet)

    def save_result(self, cal_method, score, row):
        file_name = self.file_name
        sheet = self.sheet

        # 如果文件不存在，创建新文件并写入数据
        if not os.path.exists(file_name):
            directory = os.path.dirname(file_name)
            # 检查目录是否存在，如果不存在则创建
            if not os.path.exists(directory):
                os.makedirs(directory)

            # 创建一个新的 DataFrame，并保存到指定工作表中
            df = pd.DataFrame({cal_method: [None] * row})
            df.loc[row - 1, cal_method] = score
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
        else:
            # 文件已存在，加载现有文件
            book = load_workbook(file_name)
            
            # 检查工作表是否存在
            if sheet in book.sheetnames:
                df_existing = pd.read_excel(file_name, sheet_name=sheet)
            else:
                # 如果工作表不存在，创建一个新的空DataFrame
                df_existing = pd.DataFrame()

            # 如果列已经存在，则追加数据
            if cal_method in df_existing.columns:
                if len(df_existing) < row:  # 扩展行数
                    df_existing = df_existing.reindex(list(range(row)))
                df_existing.loc[row - 1, cal_method] = score
            else:
                # 如果列不存在，创建新列并填充数据
                df_existing[cal_method] = [None] * len(df_existing)
                if len(df_existing) < row:  # 同样检查是否需要扩展行数
                    df_existing = df_existing.reindex(list(range(row)))
                df_existing.loc[row - 1, cal_method] = score
            
            # 使用 openpyxl 引擎以附加模式写入工作簿，而不会覆盖其他工作表
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # writer.book = book  # 使用已加载的工作簿
                df_existing.to_excel(writer, sheet_name=sheet, index=False)

            # 重新打开文件，设置单元格格式为数字类型，避免科学计数法
            book = load_workbook(file_name)
            sheet_to_format = book[sheet]
            
            # 设置单元格格式为浮点型数字显示
            for row_cells in sheet_to_format.iter_rows(min_row=2, max_row=row+1, min_col=1, max_col=len(df_existing.columns)):
                for cell in row_cells:
                    if isinstance(cell.value, float):  # 只对小数类型的单元格进行格式设置
                        cell.number_format = '0.0000000000000000'  # 设置足够的位数显示小数点后的所有位数

            # 保存工作簿
            book.save(file_name)
    
    # 打印并保存计算结果的函数
    def print_and_save(self, cal_method, score, row):
        # 打印结果
        print(f"\t {cal_method}: {score}")
        
        # 调用函数保存数据到指定工作表中
        self.save_to_excel(cal_method, score, row + 1)


class JsonlResultSaver(BaseResultSaver):
    """
    将数据以 JSON Lines 方式追加写入，每条记录占一行。
    """
    def __init__(self, file_path, model1, model2):
        super().__init__(file_path, model1=model1, model2=model2)

    def print_and_save(self, metrics_dict, row):
        print(f"Metrics for row={row}:")
        for k, v in metrics_dict.items():
            print(f"\t{k} = {v}")

        # 组装一个 record（一条JSON）
        record = {
            "model1": self.m1,
            "model2": self.m2,
            "row": row,
        }
        record.update(metrics_dict)  # 把 metrics_dict 的KV也加入

        # 追加写入
        directory = os.path.dirname(self.file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        with open(self.file_path, 'a', encoding='utf-8') as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


class ParquetResultSaver(BaseResultSaver):
    """
    将 (model1, model2) 在构造函数中规范化，
    然后每次将结果追加到同一个 Parquet 文件。
    """
    def __init__(self, file_path, model1, model2):
        super().__init__(file_path, model1=model1, model2=model2)

    def print_and_save(self, metrics_dict, task, language, layer):
        """
        metrics_dict: 例如 {
            "RSMNormDiff": 0.123,
            "RSA": 0.456,
            "CKA": 0.789,
            ...
        }
        layer: 用于标识的层号或某个ID
        """
        # 1) 打印所有指标
        # print(f"Metrics for layer={layer}:")
        # for k, v in metrics_dict.items():
        #     print(f"\t{k} = {v}")

        # 2) 构建 DataFrame（仅一行），包含 model1, model2, layer, 以及各指标
        data_dict = {
            "model1": [self.m1],
            "model2": [self.m2],
            "task": [task],
            "lang": [language],
            "layer": [layer],
        }
        # 把 metrics_dict 的每个键值也放入 data_dict
        for metric_name, metric_value in metrics_dict.items():
            data_dict[metric_name] = [metric_value]

        df_new = pd.DataFrame(data_dict)

        # 3) 如果文件已存在，则读出来和 df_new 做 concat
        if os.path.exists(self.file_path):
            df_existing = pd.read_parquet(self.file_path)
        else:
            df_existing = pd.DataFrame()

        # 这里可以选择是否做重复检查：如果 (m1,m2,row) 已经存在就更新或跳过
        # 简单示例：跳过写入
        # is_dup = (
        #     (df_existing["model1"] == self.m1) & 
        #     (df_existing["model2"] == self.m2) & 
        #     (df_existing["row"] == row)
        # )
        # if is_dup.any():
        #     print("Duplicate found, skip writing.")
        #     return

        # 合并后写回
        df_updated = pd.concat([df_existing, df_new], ignore_index=True)
        df_updated.to_parquet(self.file_path, index=False)




def combine_names(name1, name2, lang, max_length=31):
    def shorten_name(name):
      # 定义名称缩写的条件
      if "codeLlama-7b" in name:
          return name.replace("codeLlama-7b", "cL7b")
      elif "dsc-7b-base-v1.5" in name:
          return name.replace("dsc-7b-base-v1.5", "dsc7b")
      elif "Qwen2.5-Coder-7B" in name:
          return name.replace("Qwen2.5-Coder-7B", "QwC7b")
      # elif "codeShell-7b" in name:
      #     return name.replace("codeShell-7b", "cSh7b")
      # 根据需要添加更多缩写规则
      return name  # 如果不符合任何条件，返回原始名称

    # 依次缩写每个名称
    name1_short = shorten_name(name1)
    name2_short = shorten_name(name2)
    # name3_short = shorten_name(name3)
    
    combined_name = f"{name1_short} vs {name2_short} {lang}"
    
    # 如果缩写后的名称仍然超长，进行截取
    if len(combined_name) > max_length:
        combined_name = lang
    
    return combined_name

# 定义模型加载函数
def load_model_and_tokenizer(model_path: str, device: torch.device):
    """
    加载预训练的模型和分词器，并将模型加载到指定设备上。
    """
    tokenizer = AutoTokenizer.from_pretrained(model_path)
    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        pad_token_id=tokenizer.pad_token_id,
        torch_dtype=torch.float32
    ).to(device)
    return model, tokenizer


def _get_ngrams(segment, max_order):
  """Extracts all n-grams upto a given maximum order from an input segment.

  Args:
    segment: text segment from which n-grams will be extracted.
    max_order: maximum length in tokens of the n-grams returned by this
        methods.

  Returns:
    The Counter containing all n-grams upto max_order in segment
    with a count of how many times each n-gram occurred.
  """
  ngram_counts = collections.Counter()
  for order in range(1, max_order + 1):
    for i in range(0, len(segment) - order + 1):
      ngram = tuple(segment[i:i+order])
      ngram_counts[ngram] += 1
  return ngram_counts

def compute_bleu(reference_corpus, translation_corpus, max_order=4,
                 smooth=False):
  """Computes BLEU score of translated segments against one or more references.

  Args:
    reference_corpus: list of lists of references for each translation. Each
        reference should be tokenized into a list of tokens.
    translation_corpus: list of translations to score. Each translation
        should be tokenized into a list of tokens.
    max_order: Maximum n-gram order to use when computing BLEU score.
    smooth: Whether or not to apply Lin et al. 2004 smoothing.

  Returns:
    3-Tuple with the BLEU score, n-gram precisions, geometric mean of n-gram
    precisions and brevity penalty.
  """
  matches_by_order = [0] * max_order
  possible_matches_by_order = [0] * max_order
  reference_length = 0
  translation_length = 0
  for (references, translation) in zip(reference_corpus,
                                       translation_corpus):
    reference_length += min(len(r) for r in references)
    translation_length += len(translation)

    merged_ref_ngram_counts = collections.Counter()
    for reference in references:
      merged_ref_ngram_counts |= _get_ngrams(reference, max_order)
    translation_ngram_counts = _get_ngrams(translation, max_order)
    overlap = translation_ngram_counts & merged_ref_ngram_counts
    for ngram in overlap:
      matches_by_order[len(ngram)-1] += overlap[ngram]
    for order in range(1, max_order+1):
      possible_matches = len(translation) - order + 1
      if possible_matches > 0:
        possible_matches_by_order[order-1] += possible_matches

  precisions = [0] * max_order
  for i in range(0, max_order):
    if smooth:
      precisions[i] = ((matches_by_order[i] + 1.) /
                       (possible_matches_by_order[i] + 1.))
    else:
      if possible_matches_by_order[i] > 0:
        precisions[i] = (float(matches_by_order[i]) /
                         possible_matches_by_order[i])
      else:
        precisions[i] = 0.0

  if min(precisions) > 0:
    p_log_sum = sum((1. / max_order) * math.log(p) for p in precisions)
    geo_mean = math.exp(p_log_sum)
  else:
    geo_mean = 0

  ratio = float(translation_length) / reference_length

  if ratio > 1.0:
    bp = 1.
  else:
    bp = math.exp(1 - 1. / ratio)

  bleu = geo_mean * bp

  return (bleu, precisions, bp, ratio, translation_length, reference_length)


if __name__ == "__main__":
    # 示例使用
  # saver = ResultSaver(file_name="/newdisk/public/wws/simMeasures/results/codellama_7b_and_7b_python/test.xlsx", sheet="Metrics")
  # pwcca_mean = 0.1575322875319305  # 这是一个示例值
  # saver.print_and_save("PWCCA similarity", pwcca_mean, row=1)

  # pwcca_mean = 0.897656342134534
  # saver.print_and_save("PWCCA similarity", pwcca_mean, row=2)
  # saver = ResultSaver(file_name="/newdisk/public/wws/simMeasures/results/codellama_7b_and_7b_python/test.xlsx", sheet="Metrics1")
  # # 示例添加另一个cal_method
  # cosine_sim = 0.789456123456789
  # saver.print_and_save("Cosine similarity", cosine_sim, row=1)

  data_file_path = "/newdisk/public/wws/Dataset/CodeSearchNet/dataset/java/test.jsonl"
  prompts = extract_prompts(data_file_path, mode='codeSummary_CSearchNet', prompt_prefix="Please describe the functionality of the method: ")
  print(prompts[1])
